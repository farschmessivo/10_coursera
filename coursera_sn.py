import argparse
import requests
import json

from lxml import etree
from openpyxl import Workbook
from bs4 import BeautifulSoup
from os.path import join
from random import sample

COURSES_XML_URL = 'https://www.coursera.org/sitemap~www~courses.xml'


def convert_soup_to_text(tag):
    return tag.text if tag else None


def get_courses_urls(xml_page):
    root = etree.fromstring(xml_page)
    urls = [url.text for url in root.iter('{*}loc')]
    return urls
    print(urls)

def get_datetime_course(soup):
    json_course = convert_soup_to_text(soup.find('script', {'type': 'application/ld+json'}))
    if json_course and 'startDate' in json_course:
        return json.loads(json_course)['hasCourseInstance'][0]['startDate']


def get_course_pages(url_list):
    pages = []
    s = requests.session()
    for url in url_list
        pages.append(s.get(url).content)
    return pages


def parse_course_info(page, course_url):
    soup = BeautifulSoup(page, 'html.parser')
    course_name = soup.find('h1', {'class': 'title'}).text
    course_lang = soup.find('div', {'class': 'language-info'}).text
    course_date = get_datetime_course(soup)
    duration = len(soup.find_all('div', {'class': 'week'}))
    average_score = convert_soup_to_text(soup.find('div', {'class': 'ratings-text bt3-visible-xs'}))
    course_info = {
        'Title': course_name,
        'Language': course_lang,
        'Start Date': course_date,
        'Duration (weeks)': duration,
        'Course Rate': average_score,
        'URL': course_url
    }
    return course_info


def output_info_to_workbook(courses_info):
    headers = ['Title', 'Language', 'Start Date', 'Duration (weeks)', 'Course Rate', 'URL']
    wb = Workbook()
    sheet = wb.active
    for num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=num).value = header
    for num, info_about_course in enumerate(courses_info, 2):
        for i, key in enumerate(headers, 1):
            sheet.cell(row=num, column=i).value = info_about_course[key]
    return wb


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-p', '--path', default='', type=str, help='Input full path to folder')
    parser.add_argument('-n', '--number', default=20, type=int, help='Input number of courses')
    args = parser.parse_args()
    courses_xml_page = requests.get(COURSES_XML_URL).content
    urls = get_courses_urls(courses_xml_page)
    sample_urls = sample(urls, args.number)
    print('Getting courses info has been started')
    course_pages = get_course_pages(sample_urls)
    courses_info = [parse_course_info(page, url) for page, url in zip(course_pages, sample_urls)]
    work_book = output_info_to_workbook(courses_info)
    work_book.save(join(args.path, 'Courses from Coursera.xlsx'))
    print('Done!')